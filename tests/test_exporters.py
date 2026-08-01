from copy import deepcopy

from openpyxl import load_workbook

from markdown_exporter import export_to_markdown
from mock_test import MOCK_DATA
from qa_agent import export_to_excel
from report_storage import _app_root, slugify_project_name


def test_export_to_excel_creates_expected_stlc_sheets(tmp_path):
    output_path = tmp_path / "qa_report.xlsx"

    export_to_excel(deepcopy(MOCK_DATA), output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "01_RTM",
        "02_Test_Plan",
        "03_Test_Cases",
        "04_Environment",
        "05_Execution",
        "06_Defect_Log",
        "07_Test_Summary",
        "08_Jira_Ready",
    ]
    assert workbook["03_Test_Cases"]["A6"].value == "ID"
    assert workbook["08_Jira_Ready"]["A6"].value == "Issue Type"


def test_markdown_export_contains_traceability_and_escapes_tables():
    data = deepcopy(MOCK_DATA)
    data["casos_prueba"][0]["titulo"] = "Login valido | dashboard"

    markdown = export_to_markdown(data)

    assert "# QA Documentation - Login de usuario" in markdown
    assert "REQ-001" in markdown
    assert "Login valido \\| dashboard" in markdown
    assert "_Plantilla para completar durante la ejecucion. No se inventan defectos._" in markdown


def test_slugify_project_name_is_safe_for_report_filenames():
    assert slugify_project_name("OrangeHRM Login QA") == "orangehrm_login_qa"
    assert slugify_project_name("  ") == "proyecto"
    assert slugify_project_name("Modulo con acentos y espacios") == "modulo_con_acentos_y_espacios"


def test_app_root_uses_executable_folder_when_frozen(monkeypatch, tmp_path):
    fake_exe = tmp_path / "QA Documentation IA Agent.exe"
    monkeypatch.setattr("sys.frozen", True, raising=False)
    monkeypatch.setattr("sys.executable", str(fake_exe))

    assert _app_root() == tmp_path
