#!/usr/bin/env python3
"""
QA Documentation IA Agent
-------------------------
Genera documentacion QA manual siguiendo STLC y la exporta a Excel.
"""

import argparse
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Falta instalar openpyxl. Corre: pip install -r requirements.txt")
    sys.exit(1)

from errors import QAAgentError
from llm_providers.factory import create_provider


REQUIRED_TOP_LEVEL_KEYS = [
    "feature",
    "resumen",
    "analisis_requerimientos",
    "plan_pruebas",
    "casos_prueba",
    "checklist_entorno",
    "reporte_ejecucion",
    "defect_log",
    "test_summary",
]

REQUIRED_CASE_KEYS = [
    "id",
    "modulo",
    "titulo",
    "tipo",
    "tecnica_diseno",
    "prioridad",
    "precondiciones",
    "datos_prueba",
    "pasos",
    "resultado_esperado",
    "resultado_obtenido",
    "estado",
    "notas",
]

REQUIRED_ANALYSIS_KEYS = [
    "objetivo",
    "dudas_ambiguedades",
    "riesgos_requerimiento",
    "rtm",
]

REQUIRED_RTM_KEYS = [
    "req_id",
    "descripcion",
    "estado_qa",
    "casos_asignados",
]

REQUIRED_TEST_PLAN_KEYS = [
    "alcance",
    "fuera_de_alcance",
    "tipos_prueba",
    "tecnicas_diseno",
    "estrategia",
    "criterios_entrada",
    "criterios_salida",
    "supuestos",
    "riesgos",
]

REQUIRED_TECHNIQUE_KEYS = [
    "tecnica",
    "aplicacion",
]

REQUIRED_ENVIRONMENT_KEYS = [
    "id",
    "item",
    "responsable",
    "estado",
    "notas",
]

REQUIRED_EXECUTION_KEYS = [
    "fecha",
    "total_planificados",
    "ejecutados",
    "pasados",
    "fallados",
    "bloqueados",
    "comentarios",
    "bloqueos",
]

REQUIRED_DEFECT_KEYS = [
    "bug_id",
    "resumen",
    "severidad",
    "prioridad",
    "estado",
    "asignado_a",
    "pasos_reproduccion",
    "resultado_esperado",
    "resultado_obtenido",
    "evidencia",
    "notas",
]

REQUIRED_SUMMARY_KEYS = [
    "resumen_ejecucion",
    "estado_bugs",
    "riesgos_residuales",
    "decision",
    "comentarios_signoff",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FILL = PatternFill(start_color="173D60", end_color="173D60", fill_type="solid")
SUBTITLE_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8FBFD", end_color="F8FBFD", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF8E6", end_color="FFF8E6", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PASSED_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
FAILED_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
INFO_FILL = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
BODY_FONT = Font(size=11)
MUTED_FONT = Font(size=9, color="666666")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_TAB_COLORS = {
    "01_RTM": "1F4E79",
    "02_Test_Plan": "5B9BD5",
    "03_Test_Cases": "70AD47",
    "04_Environment": "FFC000",
    "05_Execution": "A5A5A5",
    "06_Defect_Log": "C00000",
    "07_Test_Summary": "7030A0",
    "08_Jira_Ready": "0052CC",
}


def generate_test_cases(requirement_text: str, provider_name: str = None, model: str = None) -> dict:
    """Genera documentacion QA parseada desde el proveedor configurado."""
    if not requirement_text or not requirement_text.strip():
        raise QAAgentError(
            "El requerimiento esta vacio. Pasa un texto con --text o un archivo "
            "con contenido en --input."
        )

    provider = create_provider(provider_name=provider_name, model=model)
    data = provider.generate_test_cases(requirement_text)

    if not isinstance(data, dict):
        raise QAAgentError("La respuesta del modelo no tiene la estructura esperada.")

    if "casos_prueba" not in data and "casos" not in data:
        raise QAAgentError(
            "La respuesta del modelo no tiene casos de prueba. Falta 'casos_prueba'."
        )

    normalized = _normalize_document(data)
    _validate_document_structure(normalized)
    return normalized


def _validate_required_keys(section_name: str, payload: dict, required_keys: list[str]) -> None:
    missing = [key for key in required_keys if key not in payload]
    if missing:
        raise QAAgentError(
            f"La seccion '{section_name}' esta incompleta. Faltan campos: "
            + ", ".join(missing)
        )

def _require_dict(data: dict, key: str, section_name: str = None) -> dict:
    value = data.get(key)
    if not isinstance(value, dict):
        readable_name = section_name or key
        raise QAAgentError(f"La seccion '{readable_name}' debe tener formato de objeto.")
    return value


def _require_list(data: dict, key: str, section_name: str = None, allow_empty: bool = True) -> list:
    value = data.get(key)
    if not isinstance(value, list):
        readable_name = section_name or key
        raise QAAgentError(f"La seccion '{readable_name}' debe tener formato de lista.")
    if not allow_empty and not value:
        readable_name = section_name or key
        raise QAAgentError(f"La seccion '{readable_name}' no puede estar vacia.")
    return value


def _validate_list_items(
    items: list,
    section_name: str,
    required_keys: list[str],
    list_fields: list[str] = None,
) -> None:
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise QAAgentError(f"El item #{index} de '{section_name}' no tiene formato valido.")
        _validate_required_keys(f"{section_name} #{index}", item, required_keys)
        for field in list_fields or []:
            if not isinstance(item.get(field), list):
                raise QAAgentError(
                    f"El campo '{field}' del item #{index} de '{section_name}' debe ser una lista."
                )


def _validate_document_structure(data: dict) -> None:
    """Valida que el JSON tenga la estructura esperada para el flujo STLC."""
    _validate_required_keys("documento", data, REQUIRED_TOP_LEVEL_KEYS)

    analysis = _require_dict(data, "analisis_requerimientos")
    _validate_required_keys("analisis_requerimientos", analysis, REQUIRED_ANALYSIS_KEYS)
    _require_list(analysis, "dudas_ambiguedades", "dudas_ambiguedades")
    _require_list(analysis, "riesgos_requerimiento", "riesgos_requerimiento")
    rtm_items = _require_list(analysis, "rtm", "rtm")
    _validate_list_items(rtm_items, "rtm", REQUIRED_RTM_KEYS, list_fields=["casos_asignados"])

    plan = _require_dict(data, "plan_pruebas")
    _validate_required_keys("plan_pruebas", plan, REQUIRED_TEST_PLAN_KEYS)
    for list_key in [
        "alcance",
        "fuera_de_alcance",
        "tipos_prueba",
        "criterios_entrada",
        "criterios_salida",
        "supuestos",
        "riesgos",
    ]:
        _require_list(plan, list_key)
    techniques = _require_list(plan, "tecnicas_diseno", "tecnicas_diseno")
    _validate_list_items(techniques, "tecnicas_diseno", REQUIRED_TECHNIQUE_KEYS)

    cases = _require_list(data, "casos_prueba", allow_empty=False)
    for index, case in enumerate(cases, start=1):
        if not isinstance(case, dict):
            raise QAAgentError(f"El caso de prueba #{index} no tiene formato valido.")
        _validate_required_keys(f"caso de prueba #{index}", case, REQUIRED_CASE_KEYS)
        if not isinstance(case.get("pasos"), list):
            case_id = case.get("id", f"#{index}")
            raise QAAgentError(f"El caso {case_id} debe tener 'pasos' como lista.")

    environment_items = _require_list(data, "checklist_entorno")
    _validate_list_items(environment_items, "checklist_entorno", REQUIRED_ENVIRONMENT_KEYS)

    execution = _require_dict(data, "reporte_ejecucion")
    _validate_required_keys("reporte_ejecucion", execution, REQUIRED_EXECUTION_KEYS)
    _require_list(execution, "bloqueos", "bloqueos")

    defects = _require_list(data, "defect_log")
    _validate_list_items(defects, "defect_log", REQUIRED_DEFECT_KEYS)

    summary = _require_dict(data, "test_summary")
    _validate_required_keys("test_summary", summary, REQUIRED_SUMMARY_KEYS)


def _normalize_document(data: dict) -> dict:
    """Adapta respuestas viejas al formato STLC actual."""
    if "casos_prueba" in data:
        return data

    alcance = data.get("alcance") or {}
    casos = data.get("casos") or []
    normalized_cases = []
    for caso in casos:
        normalized_cases.append(
            {
                "id": caso.get("id", ""),
                "modulo": data.get("feature", ""),
                "titulo": caso.get("titulo", ""),
                "tipo": caso.get("tipo", ""),
                "prioridad": caso.get("prioridad", ""),
                "precondiciones": caso.get("precondiciones", ""),
                "datos_prueba": "",
                "tecnica_diseno": "",
                "pasos": caso.get("pasos", []),
                "resultado_esperado": caso.get("resultado_esperado", ""),
                "resultado_obtenido": "",
                "estado": "Pendiente",
                "notas": caso.get("notas", ""),
            }
        )

    return {
        "feature": data.get("feature", "Feature"),
        "resumen": data.get("resumen", ""),
        "analisis_requerimientos": {
            "objetivo": data.get("resumen", ""),
            "dudas_ambiguedades": [],
            "riesgos_requerimiento": data.get("supuestos_y_riesgos", []),
            "rtm": [],
        },
        "plan_pruebas": {
            "alcance": alcance.get("incluye", []),
            "fuera_de_alcance": alcance.get("no_incluye", []),
            "tipos_prueba": ["Funcional manual"],
            "tecnicas_diseno": [],
            "estrategia": data.get("estrategia", ""),
            "criterios_entrada": [],
            "criterios_salida": [],
            "supuestos": data.get("supuestos_y_riesgos", []),
            "riesgos": [],
        },
        "casos_prueba": normalized_cases,
        "checklist_entorno": [],
        "reporte_ejecucion": _empty_execution_report(),
        "defect_log": [_empty_defect()],
        "test_summary": _empty_summary(),
    }


def _empty_execution_report() -> dict:
    return {
        "fecha": "",
        "total_planificados": "",
        "ejecutados": "",
        "pasados": "",
        "fallados": "",
        "bloqueados": "",
        "comentarios": "",
        "bloqueos": [],
    }


def _empty_defect() -> dict:
    return {
        "bug_id": "",
        "resumen": "",
        "severidad": "",
        "prioridad": "",
        "estado": "",
        "asignado_a": "",
        "pasos_reproduccion": "",
        "resultado_esperado": "",
        "resultado_obtenido": "",
        "evidencia": "",
        "notas": "",
    }


def _empty_summary() -> dict:
    return {
        "resumen_ejecucion": "",
        "estado_bugs": "",
        "riesgos_residuales": "",
        "decision": "Pendiente de ejecucion",
        "comentarios_signoff": "",
    }


def _setup_sheet(ws, title: str, data: dict) -> int:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.sheet_properties.tabColor = SHEET_TAB_COLORS.get(ws.title, "1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.cell(row=2, column=1, value=f"Feature: {data.get('feature', 'N/A')}").font = BODY_FONT
    ws.cell(row=3, column=1, value=data.get("resumen", "")).font = MUTED_FONT
    ws.cell(row=4, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = MUTED_FONT
    return 6


def _write_section(ws, row: int, title: str, value) -> int:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SUBTITLE_FILL
    ws.cell(row=row, column=1).border = BORDER
    row += 1
    if isinstance(value, list):
        formatted_items = []
        for item in value:
            if isinstance(item, dict):
                technique = item.get("tecnica") or item.get("nombre") or ""
                application = item.get("aplicacion") or item.get("detalle") or ""
                formatted_items.append(f"- {technique}: {application}".strip())
            else:
                formatted_items.append(f"- {item}")
        text = "\n".join(formatted_items) if formatted_items else ""
    else:
        text = value or ""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = BODY_FONT
    cell.alignment = WRAP
    cell.border = BORDER
    ws.row_dimensions[row].height = max(24, min(96, 18 * (str(text).count("\n") + 1)))
    row += 2
    return row


def _write_table(ws, row: int, headers: list[str], rows: list[list], widths: list[int]) -> int:
    start_row = row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 30

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    row += 1
    if not rows:
        rows = [["" for _ in headers]]

    for row_values in rows:
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL
            _apply_value_style(cell, headers[col - 1])
        ws.row_dimensions[row].height = _estimate_row_height(row_values)
        row += 1

    end_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start_row}:{end_col}{row - 1}"
    ws.freeze_panes = f"A{start_row + 1}"
    return row + 1


def _estimate_row_height(row_values: list) -> int:
    max_lines = 1
    for value in row_values:
        line_count = str(value or "").count("\n") + 1
        max_lines = max(max_lines, line_count)
    return max(24, min(120, 18 * max_lines))


def _apply_value_style(cell, header: str) -> None:
    value = str(cell.value or "").strip().lower()
    header_lower = header.lower()

    if header_lower in {"resultado obtenido", "notas", "evidencia", "comentarios"}:
        cell.fill = INPUT_FILL

    if header_lower in {"estado", "estado qa", "prioridad", "severidad"}:
        cell.alignment = CENTER_WRAP

    if value in {"pendiente", "en revision", "duda abierta"}:
        cell.fill = PENDING_FILL
    elif value in {"revisado", "pasado", "resuelto"}:
        cell.fill = PASSED_FILL
    elif value in {"fallado", "critica", "critico", "alta"}:
        cell.fill = FAILED_FILL
    elif value in {"media", "baja", "bloqueado"}:
        cell.fill = INFO_FILL


def _add_dropdown(ws, cell_range: str, options: list[str]) -> None:
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    validation.add(cell_range)


def _steps_text(steps: list[str]) -> str:
    return "\n".join(f"{index}. {step}" for index, step in enumerate(steps or [], start=1))


def _write_rtm_sheet(wb, data: dict):
    ws = wb.active
    ws.title = "01_RTM"
    row = _setup_sheet(ws, "01 - Analisis de Requerimientos / RTM", data)
    analysis = data.get("analisis_requerimientos") or {}
    row = _write_section(ws, row, "Objetivo QA", analysis.get("objetivo", ""))
    row = _write_section(ws, row, "Dudas y ambiguedades", analysis.get("dudas_ambiguedades", []))
    row = _write_section(ws, row, "Riesgos del requerimiento", analysis.get("riesgos_requerimiento", []))

    headers = ["Req ID", "Descripcion del Requerimiento", "Estado QA", "ID Caso de Prueba Asignado"]
    table_rows = [
        [
            item.get("req_id", ""),
            item.get("descripcion", ""),
            item.get("estado_qa", ""),
            ", ".join(item.get("casos_asignados") or []),
        ]
        for item in analysis.get("rtm", [])
    ]
    _write_table(ws, row, headers, table_rows, [14, 60, 18, 30])
    _add_dropdown(ws, f"C{row + 1}:C200", ["Revisado", "En revision", "Duda abierta"])


def _write_test_plan_sheet(wb, data: dict):
    ws = wb.create_sheet("02_Test_Plan")
    row = _setup_sheet(ws, "02 - Plan de Pruebas", data)
    plan = data.get("plan_pruebas") or {}
    for title, key in [
        ("Alcance (In-Scope)", "alcance"),
        ("Fuera de Alcance (Out-of-Scope)", "fuera_de_alcance"),
        ("Tipos de prueba", "tipos_prueba"),
        ("Tecnicas de diseno sugeridas (ISTQB)", "tecnicas_diseno"),
        ("Estrategia", "estrategia"),
        ("Criterios de entrada", "criterios_entrada"),
        ("Criterios de salida", "criterios_salida"),
        ("Supuestos", "supuestos"),
        ("Riesgos", "riesgos"),
    ]:
        row = _write_section(ws, row, title, plan.get(key, [] if key != "estrategia" else ""))
    ws.column_dimensions["A"].width = 110


def _write_test_cases_sheet(wb, data: dict):
    ws = wb.create_sheet("03_Test_Cases")
    row = _setup_sheet(ws, "03 - Casos de Prueba", data)
    headers = [
        "ID", "Modulo", "Titulo", "Tipo", "Tecnica de diseno", "Prioridad", "Precondiciones",
        "Datos de prueba", "Pasos", "Resultado esperado", "Resultado obtenido",
        "Estado", "Notas",
    ]
    rows = []
    for case in data.get("casos_prueba", []):
        rows.append(
            [
                case.get("id", ""),
                case.get("modulo", ""),
                case.get("titulo", ""),
                case.get("tipo", ""),
                case.get("tecnica_diseno", ""),
                case.get("prioridad", ""),
                case.get("precondiciones", ""),
                case.get("datos_prueba", ""),
                _steps_text(case.get("pasos", [])),
                case.get("resultado_esperado", ""),
                case.get("resultado_obtenido", ""),
                case.get("estado", "Pendiente"),
                case.get("notas", ""),
            ]
        )
    _write_table(ws, row, headers, rows, [10, 18, 32, 14, 26, 14, 30, 28, 46, 34, 34, 14, 26])
    _add_dropdown(ws, f"F{row + 1}:F300", ["Alta", "Media", "Baja"])
    _add_dropdown(ws, f"L{row + 1}:L300", ["Pendiente", "Pasado", "Fallado", "Bloqueado", "No ejecutado"])


def _write_environment_sheet(wb, data: dict):
    ws = wb.create_sheet("04_Environment")
    row = _setup_sheet(ws, "04 - Checklist de Entorno", data)
    headers = ["ID", "Item", "Responsable", "Estado", "Notas"]
    rows = [
        [
            item.get("id", ""),
            item.get("item", ""),
            item.get("responsable", ""),
            item.get("estado", "Pendiente"),
            item.get("notas", ""),
        ]
        for item in data.get("checklist_entorno", [])
    ]
    _write_table(ws, row, headers, rows, [12, 70, 18, 16, 34])
    _add_dropdown(ws, f"D{row + 1}:D150", ["Pendiente", "Listo", "Bloqueado", "No aplica"])


def _write_execution_sheet(wb, data: dict):
    ws = wb.create_sheet("05_Execution")
    row = _setup_sheet(ws, "05 - Reporte Diario de Ejecucion", data)
    report = data.get("reporte_ejecucion") or _empty_execution_report()
    headers = ["Fecha", "Total planificados", "Ejecutados", "Pasados", "Fallados", "Bloqueados", "Comentarios"]
    rows = [[
        report.get("fecha", ""),
        report.get("total_planificados", ""),
        report.get("ejecutados", ""),
        report.get("pasados", ""),
        report.get("fallados", ""),
        report.get("bloqueados", ""),
        report.get("comentarios", ""),
    ]]
    row = _write_table(ws, row, headers, rows, [16, 20, 16, 16, 16, 16, 50])
    _write_section(ws, row, "Bloqueos / pendientes", report.get("bloqueos", []))


def _write_defect_log_sheet(wb, data: dict):
    ws = wb.create_sheet("06_Defect_Log")
    row = _setup_sheet(ws, "06 - Registro de Defectos", data)
    headers = [
        "Bug ID", "Resumen", "Severidad", "Prioridad", "Estado", "Asignado a",
        "Pasos reproduccion", "Resultado esperado", "Resultado obtenido", "Evidencia", "Notas",
    ]
    defects = data.get("defect_log") or [_empty_defect()]
    rows = [
        [
            defect.get("bug_id", ""),
            defect.get("resumen", ""),
            defect.get("severidad", ""),
            defect.get("prioridad", ""),
            defect.get("estado", ""),
            defect.get("asignado_a", ""),
            defect.get("pasos_reproduccion", ""),
            defect.get("resultado_esperado", ""),
            defect.get("resultado_obtenido", ""),
            defect.get("evidencia", ""),
            defect.get("notas", ""),
        ]
        for defect in defects
    ]
    _write_table(ws, row, headers, rows, [12, 34, 14, 14, 16, 18, 44, 34, 34, 24, 24])
    _add_dropdown(ws, f"C{row + 1}:C200", ["Critica", "Alta", "Media", "Baja"])
    _add_dropdown(ws, f"D{row + 1}:D200", ["Alta", "Media", "Baja"])
    _add_dropdown(ws, f"E{row + 1}:E200", ["Abierto", "En progreso", "Resuelto", "Cerrado", "Bloqueado"])


def _write_summary_sheet(wb, data: dict):
    ws = wb.create_sheet("07_Test_Summary")
    row = _setup_sheet(ws, "07 - Test Summary / Sign-off", data)
    summary = data.get("test_summary") or _empty_summary()
    for title, key in [
        ("Resumen de ejecucion", "resumen_ejecucion"),
        ("Estado de bugs", "estado_bugs"),
        ("Riesgos residuales", "riesgos_residuales"),
        ("Decision / Sign-off", "decision"),
        ("Comentarios de cierre", "comentarios_signoff"),
    ]:
        row = _write_section(ws, row, title, summary.get(key, ""))
    ws.column_dimensions["A"].width = 110


def _requirement_ids_for_case(data: dict, case_id: str) -> str:
    analysis = data.get("analisis_requerimientos") or {}
    req_ids = []
    for item in analysis.get("rtm", []):
        if case_id in (item.get("casos_asignados") or []):
            req_ids.append(item.get("req_id", ""))
    return ", ".join(req_id for req_id in req_ids if req_id)


def _write_jira_ready_sheet(wb, data: dict):
    ws = wb.create_sheet("08_Jira_Ready")
    row = _setup_sheet(ws, "08 - Jira Ready", data)
    headers = [
        "Issue Type", "Summary", "Description", "Priority", "Labels", "Component",
        "Requirement IDs", "Test Case ID", "Test Steps", "Expected Result", "Status",
    ]
    rows = []
    for case in data.get("casos_prueba", []):
        case_id = case.get("id", "")
        labels = [
            "qa-documentation",
            "manual-test",
            str(case.get("tipo", "")).lower().replace(" ", "-"),
            str(case.get("tecnica_diseno", "")).lower().replace(" ", "-"),
        ]
        rows.append(
            [
                "Test",
                f"{case_id} - {case.get('titulo', '')}",
                (
                    f"Modulo: {case.get('modulo', '')}\n"
                    f"Precondiciones: {case.get('precondiciones', '')}\n"
                    f"Datos de prueba: {case.get('datos_prueba', '')}\n"
                    f"Tecnica de diseno: {case.get('tecnica_diseno', '')}"
                ),
                case.get("prioridad", ""),
                ", ".join(label for label in labels if label and label != "-"),
                case.get("modulo", ""),
                _requirement_ids_for_case(data, case_id),
                case_id,
                _steps_text(case.get("pasos", [])),
                case.get("resultado_esperado", ""),
                case.get("estado", "Pendiente"),
            ]
        )
    _write_table(ws, row, headers, rows, [14, 44, 54, 14, 38, 20, 20, 16, 48, 42, 16])
    _add_dropdown(ws, f"A{row + 1}:A300", ["Test", "Task", "Story"])
    _add_dropdown(ws, f"D{row + 1}:D300", ["Alta", "Media", "Baja"])
    _add_dropdown(ws, f"K{row + 1}:K300", ["Pendiente", "Pasado", "Fallado", "Bloqueado", "No ejecutado"])


def export_to_excel(data: dict, output_path: str):
    """Exporta la documentacion STLC a un Excel listo para revisar y completar."""
    data = _normalize_document(data)
    wb = Workbook()
    _write_rtm_sheet(wb, data)
    _write_test_plan_sheet(wb, data)
    _write_test_cases_sheet(wb, data)
    _write_environment_sheet(wb, data)
    _write_execution_sheet(wb, data)
    _write_defect_log_sheet(wb, data)
    _write_summary_sheet(wb, data)
    _write_jira_ready_sheet(wb, data)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Agente de IA que genera documentacion QA siguiendo STLC."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--input", help="Ruta a un archivo .txt/.md con el requerimiento")
    group.add_argument("--text", help="El requerimiento pasado directo como texto")
    parser.add_argument("--output", default="documentacion_qa.xlsx", help="Ruta del Excel de salida")
    parser.add_argument("--provider", default=None, help="Proveedor de IA a usar: gemini")
    parser.add_argument("--model", default=None, help="Modelo especifico a usar")
    args = parser.parse_args()

    if args.input:
        if not os.path.exists(args.input):
            print(f"Error: no se encontro el archivo: {args.input}")
            sys.exit(1)
        with open(args.input, "r", encoding="utf-8") as f:
            requirement_text = f.read()
    else:
        requirement_text = args.text

    print("Generando documentacion QA STLC con IA...")
    try:
        data = generate_test_cases(requirement_text, provider_name=args.provider, model=args.model)
    except QAAgentError as e:
        print(f"Error: {e}")
        sys.exit(1)

    n_cases = len(data.get("casos_prueba", []))
    print(f"OK: se generaron {n_cases} casos de prueba y plantillas STLC.")

    try:
        export_to_excel(data, args.output)
    except Exception as e:
        print(f"Error: no se pudo generar el Excel: {e}")
        sys.exit(1)

    print(f"Excel guardado en: {args.output}")


if __name__ == "__main__":
    main()
